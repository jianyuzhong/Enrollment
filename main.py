from Spider import Spider
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def main():
    print("Hello World")
def set_excel_style(output_path:str):
    wb = load_workbook(output_path)
    ws = wb.active
    ws.column_dimensions['B'].width=40
    ws.column_dimensions['C'].width=40
    ws.column_dimensions['D'].width=100
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

    # 保存修改后的文件
    wb.save(output_path)
def set_excel_style1(output_path:str):
    wb = load_workbook(output_path)
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        cell = ws[f'C{row}']
        if row==10:
            a=1
        if cell.value: 
            cell.value=cell.value.replace('\n\n\n', '').replace('      ', ' ') # 确保单元格有值
            cell.value = repmoven( cell.value)

    # 保存修改后的 Excel 文件
    wb.save('example_modified.xlsx')
def set_excel_style2(output_path:str):
    wb = load_workbook(output_path)  # 替换为你的文件名
    # wb = openpyxl.Workbook()  # 如果你想要创建新的工作簿

    # 选择工作表
    ws = wb.active  # 替换为你的工作表名，如 ws = wb['Sheet1']

    # 遍历所有单元格并设置对齐方式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center',wrapText=True)  # 垂直居中
    wb.save('example_modified1.xlsx') 
def repmoven(input_str:str):
    # 分割字符串
    parts = input_str.split('\n(\n')

    # 处理 parts[1] 去掉所有 \n
    if len(parts) > 1:
        parts[1] = parts[1].replace('\n', ' ')
    else:
        return input_str
    # 重新组合字符串
    output_str = f"{parts[0]}\n({parts[1]}"
    return output_str

if __name__=="__main__":
    spider = Spider()
    # spider.get_html_content(r"C:\Users\jianyuzhong\Desktop\data\2023年成都市青羊区小学毕业生初中入学划片范围.html",r"C:\Users\jianyuzhong\Desktop\data\2023年成都市青羊区小学划片范围一览表.html")
    # spider.get_html_content(r"C:\Users\jianyuzhong\Desktop\data\2023年成都高新区小学毕业生初中入学划片范围.html",r"C:\Users\jianyuzhong\Desktop\data\2023年成都市高新区小学划片范围一览表.html")
    #spider.get_html_content(r"C:\Users\jianyuzhong\Desktop\data\2023年成都市锦江区小学毕业生初中入学划片范围.html",r"C:\Users\jianyuzhong\Desktop\data\2023年成都市锦江区小学划片范围一览表.html")
    # spider.get_html2_content(r"C:\Users\jianyuzhong\Desktop\data\2024年成都市青羊区小学划片范围一览表.html")
    # set_excel_style("enrollment_data.xlsx")
    # set_excel_style1("高新区.xlsx")
    set_excel_style2("example_modified.xlsx")

